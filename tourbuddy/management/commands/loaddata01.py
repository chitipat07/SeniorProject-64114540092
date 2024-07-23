from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from tourbuddy.models import Trip

class Command(BaseCommand):
    help = "Load website data from สถานที่ท่องเที่ยวver.แก้ไข.xlsx."

    def handle(self, *args, **options):
        file_path = 'สถานที่ท่องเที่ยวNEW.xlsx'  

        wb = load_workbook(file_path)
        ws = wb['Trip']

        for row in ws.iter_rows(min_row=2, values_only=True):  
            location, latitude, longitude, name, rating, review, trip_type,g_type, image = row

            if not location:  
                continue

            
            trip = Trip(
                location=location,
                latitude=float(latitude) if latitude else None,
                longitude=float(longitude) if longitude else None,
                name=name,
                rating=float(rating) if rating else None,
                review=review,
                trip_type=trip_type,
                g_type=g_type,
                image=image,
            )
            trip.save()